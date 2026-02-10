import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# Utility: Data Quality Check
# ============================================================
def data_quality_check(df, name="Dataset"):
    print(f"\n{'='*60}")
    print(f"  Data Quality Report: {name}")
    print(f"{'='*60}")
    print(f"Shape: {df.shape}\n")

    print("Columns:")
    for col in df.columns:
        print(f"  - {col}")

    nulls = df.isnull().sum()
    print(f"\nNull counts:")
    for col, count in nulls.items():
        pct = count / len(df) * 100
        print(f"  {col}: {count} ({pct:.1f}%)")

    print(f"\nEmpty/whitespace-only strings:")
    for col in df.select_dtypes(include="object").columns:
        empty = (df[col].astype(str).str.strip() == "").sum()
        if empty > 0:
            print(f"  {col}: {empty}")

    dupes = df.duplicated().sum()
    print(f"\nDuplicate rows: {dupes}")

    print(f"\nColumn dtypes:")
    for col, dtype in df.dtypes.items():
        print(f"  {col}: {dtype}")

    print(f"\n{'='*60}\n")


# ============================================================
# Utility: Save DataFrame as formatted Excel
# ============================================================
def save_formatted_xlsx(df, filepath, sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # -- Styles --
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    # -- Write headers --
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # -- Write data rows --
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            # Convert pandas NaT/NaN to None so Excel shows blank
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # -- Auto-fit column widths --
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):  # Sample first 100 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # -- Freeze header row --
    ws.freeze_panes = "A2"

    # -- Auto-filter --
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"Saved: {filepath}")


# ============================================================
# 1. Clean NPS Data
# ============================================================
def clean_nps(path="data/raw/nps_data.xlsx"):
    df = pd.read_excel(path)

    # Drop any unnamed/ghost columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    # Drop fully empty rows
    df = df.dropna(how="all")

    # Drop duplicates
    df = df.drop_duplicates()

    # Strip whitespace from string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # Replace empty strings with NaN
    df = df.replace("", pd.NA)

    # Parse date column
    df["wootricsnps_responses.created_at_date"] = pd.to_datetime(
        df["wootricsnps_responses.created_at_date"], errors="coerce"
    )

    # Ensure score is numeric
    df["wootricsnps_responses.score"] = pd.to_numeric(
        df["wootricsnps_responses.score"], errors="coerce"
    )

    # Standardize column names
    df.columns = (
        df.columns.str.strip()
        .str.lower()
        .str.replace(".", "_", regex=False)
        .str.replace(" ", "_", regex=False)
    )

    print(f"NPS after cleanup: {df.shape}")
    return df


# ============================================================
# 2. Clean CSAT Data
# ============================================================
def clean_csat(path="data/raw/csat.xlsx"):
    df = pd.read_excel(path)

    # Drop any unnamed/ghost columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    # Drop fully empty rows
    df = df.dropna(how="all")

    # Drop duplicates
    df = df.drop_duplicates()

    # Strip whitespace from string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # Replace empty strings with NaN
    df = df.replace("", pd.NA)

    # Standardize column names
    df.columns = (
        df.columns.str.strip()
        .str.lower()
        .str.replace(".", "_", regex=False)
        .str.replace(" ", "_", regex=False)
    )

    print(f"CSAT after cleanup: {df.shape}")
    return df


# ============================================================
# 3. Clean Support Tickets
# ============================================================
def clean_support_tickets(path="data/raw/support_tickets.xlsx"):
    # Only read the 14 real columns — ignore 317 ghost columns
    df = pd.read_excel(path, usecols=range(14))

    # Drop rows that are entirely empty
    df = df.dropna(how="all")

    # Drop duplicates
    df = df.drop_duplicates()

    # Strip whitespace from string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # Replace empty strings with NaN
    df = df.replace("", pd.NA)

    # Parse date columns
    for col in ["CLOSED_AT", "CREATED_AT", "CSAT_SUBMITTED_AT"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    # Ensure numeric columns are numeric
    for col in ["PORTAL_ID", "TICKET_ID"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Standardize column names
    df.columns = (
        df.columns.str.strip()
        .str.lower()
        .str.replace(".", "_", regex=False)
        .str.replace(" ", "_", regex=False)
    )

    print(f"Support tickets after cleanup: {df.shape}")
    return df


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":

    # ---- Step 1: Raw data quality check ----
    print("=" * 60)
    print("  RAW DATA QUALITY CHECK")
    print("=" * 60)

    raw_dir = Path("data/raw")
    for file in raw_dir.glob("*.xlsx"):
        print(f"\nLoading {file.name}...")
        if file.stem == "support_tickets":
            df = pd.read_excel(file, usecols=range(14))
        else:
            df = pd.read_excel(file)
        data_quality_check(df, file.stem)

    # ---- Step 2: Clean each dataset ----
    print("=" * 60)
    print("  CLEANING DATA...")
    print("=" * 60)

    nps = clean_nps()
    csat = clean_csat()
    tickets = clean_support_tickets()

    # ---- Step 3: Save as formatted xlsx ----
    processed_dir = Path("data/processed")
    processed_dir.mkdir(parents=True, exist_ok=True)

 #   save_formatted_xlsx(nps, processed_dir / "nps_clean.xlsx", "NPS Data")
  #  save_formatted_xlsx(csat, processed_dir / "csat_clean.xlsx", "CSAT Data")
   # save_formatted_xlsx(tickets, processed_dir / "support_tickets_clean.xlsx", "Support Tickets")

    # ---- Step 4: Post-cleaning quality check ----
    print("\n" + "=" * 60)
    print("  POST-CLEANING QUALITY CHECK")
    print("=" * 60)

    data_quality_check(nps, "NPS (cleaned)")
    data_quality_check(csat, "CSAT (cleaned)")
    data_quality_check(tickets, "Support Tickets (cleaned)")

    print("Done! Cleaned files saved to data/processed/") 


print("Unique portal_ids in CSAT:", csat["wootricscsat_responses_portal_id"].nunique())
print("Unique portal_ids in Tickets:", tickets["portal_id"].nunique())

# Check how many match
common = set(csat["wootricscsat_responses_portal_id"].dropna()) & set(tickets["portal_id"].dropna())
print(f"Matching portal_ids: {len(common)}")